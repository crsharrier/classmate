from django.core.management.base import BaseCommand
from django.utils import timezone
from django.db.utils import IntegrityError
from datetime import date
from classmate.models import Class, Student, ClassAssignment, TermPeriod

from openpyxl import load_workbook
import pandas as pd
from datetime import datetime, date
from classmate.config import HOMEWORK_TRACKER_PATH

def populate_classes():
    classes = ['Whole Class', 'Maths', 'Reading', 'Writing']
    for cl in classes:
        Class.objects.create(class_name=cl)

def populate_students():
    wb = load_workbook(HOMEWORK_TRACKER_PATH)
    period = TermPeriod.objects.get(week_commencing=date(2024, 9, 2))
    new_students = []
    class_assignments = []

    for class_name in ('Reading', 'Writing', 'Maths', 'Whole Class'):
        cls = Class.objects.get(class_name=class_name)
        ws = wb['Certificates' if class_name == 'Whole Class' else class_name]

        row_index = 2
        current_name = ws['A' + str(row_index)].value
        current_name = current_name.strip() if isinstance(current_name, str) else current_name

        while current_name:
            try:
                student = Student.objects.get(student_name=current_name)
            except Student.DoesNotExist:
                student = Student(student_name=current_name)
                new_students.append(student)

            ca = ClassAssignment(student=student, 
                            the_class=cls, 
                            week_commencing=period)
            class_assignments.append(ca)

            row_index += 1
            current_name = ws['A' + str(row_index)].value
            current_name = current_name.strip() if isinstance(current_name, str) else current_name

    for st in new_students:
        try:
            st.save()
        except IntegrityError:
            continue

    for ca in class_assignments:
        try:
            ca.save()
        except IntegrityError:
            continue


def populate_term_periods():
    df = pd.read_csv('data/term_periods.csv')
    for _, row in df.iterrows():
        wc = datetime.strptime(row['week_commencing'], '%d/%m/%y')

        tp = TermPeriod(
            week_commencing = wc,
            week_name = row['week_name'],
            period_name = row['period_name'],
            period_type = row['period_type'],
        )
        tp.save()

class Command(BaseCommand):
    help = 'Populate db with data'

    def add_arguments(self, parser):
        parser.add_argument(
            '--all',
            action='store_true',
            help="Populate all models with data"
        )

        parser.add_argument(
            '--classes',
            action='store_true',
            help="Populate 'Class' model with data"
        )

        parser.add_argument(
            '--students',
            action='store_true',
            help="Populate 'Student' model with data"
        )

        parser.add_argument(
            '--term-periods',
            action='store_true',
            help="Populate 'ClassAssignment' model with data"
        )

        # parser.add_argument(
        #     '--dates',
        #     action='store_true',
        #     help="Populate 'ClassAssignment' model with data"
        # )

    def handle(self, *args, **options):
        if options.get('all'):
            populate_term_periods()
            populate_classes()
            populate_students()
            self.stdout.write(self.style.SUCCESS("All entities populated successfully"))

        if options.get('classes'):
            populate_classes()
            self.stdout.write(self.style.SUCCESS("'Class' populated successfully"))

        if options.get('students'):
            populate_students()
            self.stdout.write(self.style.SUCCESS("'Student' populated successfully"))

        if options.get('term_periods'):
            populate_term_periods()
            self.stdout.write(self.style.SUCCESS("'TermPeriods' populated successfully"))

        # if options.get('dates'):
        #     populate_dates()
        #     self.stdout.write(self.style.SUCCESS("'DateDim' populated successfully"))

        if not any(options.get(key) for key in ['all', 'classes', 'students', 'term_periods']):
            self.stdout.write(self.style.WARNING(
                "No arguments provided. Please use one of the following options:\n"
                "--classes to populate 'Class' model\n"
                "--students to populate 'Student' model\n"
                "--class-assignments to populate 'ClassAssignment' model"
            ))