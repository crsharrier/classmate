from classmate.models import Student, Class, Job, SeatingAssignment, ClassAssignment, TermPeriod
from classmate.config import *
from classmate.config import (
    SeatingPlanCellReferences as ST, LiningUpCellReferences as LN)

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import random


#TODO: Consider past seating
def generate_seating_plan(the_class: Class):
    '''
    Generate a list of 30 (Student | None). 
    Nulls are used to designate empty seats.
    '''
    class_assigments = ClassAssignment.objects.filter(the_class=the_class,
                                                      is_current=True)
    
    students = [ca.student for ca in class_assigments]

    random.shuffle(students)
    if len(students) < NUM_DESKS:
        students += [None] * (NUM_DESKS - len(students))
    
    desks = []
    for desk_name, desk in ST.desks.items():
        for seat_num, cell_ref in desk:
            desks.append((desk_name, seat_num, cell_ref))

    desk_details = [(student, desk_name, seat_num, cell_ref)
                    for student, (desk_name, seat_num, cell_ref) in 
                    zip(students, desks)]

    return desk_details


def print_seating_plans(wb: Workbook,
                        classes: dict[Class, list[tuple[Class, Student, str, int, str]]],
                        term_period: TermPeriod):
    
    ws = wb['Seating Plans']

    # print header:
    ws[ST.commencing] = term_period.week_commencing
    ws[ST.period] = term_period.period_name

    # print students:
    for i, (cls, details) in enumerate(classes.items()):
        ws[ST.titles[i]] = cls.class_name
        for (student, desk_name, seat_num, orig_cell_ref) in details:

            col_diff = ws[ST.titles[i]].column - ws[ST.titles[0]].column
            row_diff = ws[ST.titles[i]].row - ws[ST.titles[0]].row
            cell = (None if orig_cell_ref is None else
                        ws.cell(ws[orig_cell_ref].row + row_diff,
                               ws[orig_cell_ref].column + col_diff))

            cell.value = '' if student is None else student.student_name


#TODO: Consider past lining up orders
def generate_lining_up_order(class_name: str):

    students = [ca.student for ca in 
                ClassAssignment.objects
                .filter(the_class__class_name=class_name,
                        is_current=True)]
    
    random.shuffle(students)
    if len(students) < NUM_DESKS:
        students += [None] * (NUM_DESKS - len(students))
    print(class_name, len([students]))
    return students


def print_lining_up_order(ws: Worksheet,
                          class_name: str,
                          students: list[Student],
                          col: int):
    
    ws.cell(LN.title_row, col, value=class_name)
    for i, st in enumerate(students):
        row = i + LN.start_row
        name = '' if st is None else st.student_name 
        ws.cell(row, col, value=name)


def print_lining_up_orders(wb: Workbook,
                          lining_up_orders: dict[str, list[Student]],
                          term_period: TermPeriod):
    ws = wb['Lining Up']

    # print header:
    ws[LN.commencing] = term_period.week_commencing
    ws[LN.period] = term_period.period_name

    for i, (class_name, students) in enumerate(lining_up_orders.items()):
        # print(i, class_name, len(students), '\n\n')
        print_lining_up_order(ws, class_name, students, LN.cols[i])
    